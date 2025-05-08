#!/usr/bin/env python3
"""
deepseek_translate.py  –  Translate .docx / .xlsx via DeepSeek, keep formatting
---------------------------------------------------------------------------
• Word / Excel 双格式      • System 提示锁定 "segments"
• 键名兜底容错             • 兼容 openai-python 0.x / 1.x
"""

import argparse, json, os, sys, time
from pathlib import Path
from typing import List, Dict

# ---------- OpenAI SDK 兼容层 --------------------------------------------
try:                  # ≥1.0
    from openai import OpenAI
    _new_sdk = True
except ImportError:   # 0.x
    import openai
    _new_sdk = False
# -------------------------------------------------------------------------

from docx import Document
from openpyxl import load_workbook
import tiktoken
from tqdm import tqdm

# ======================== CONFIG ========================================
API_KEY  = os.getenv("DEEPSEEK_API_KEY", "sk-4b9897e660ec4208af191aed958da109")
BASE_URL = "https://api.deepseek.com"
MODEL    = "deepseek-chat"
MAX_BATCH_TOKENS = 3000
TEMPERATURE      = 0.2
REQUEST_TIMEOUT  = 120

# 语言代码映射表
LANGUAGE_CODES = {
    # 东亚语言
    "zh": "Chinese (简体中文)",
    "zh-tw": "Traditional Chinese (繁體中文)",
    "ja": "Japanese (日本語)",
    "ko": "Korean (한국어)",
    
    # 欧洲语言
    "en": "English",
    "fr": "French (Français)",
    "de": "German (Deutsch)",
    "es": "Spanish (Español)",
    "it": "Italian (Italiano)",
    "pt": "Portuguese (Português)",
    "ru": "Russian (Русский)",
    "nl": "Dutch (Nederlands)",
    "pl": "Polish (Polski)",
    "tr": "Turkish (Türkçe)",
    
    # 东南亚语言
    "vi": "Vietnamese (Tiếng Việt)",
    "th": "Thai (ไทย)",
    "id": "Indonesian (Bahasa Indonesia)",
    "ms": "Malay (Bahasa Melayu)",
    
    # 南亚语言
    "hi": "Hindi (हिन्दी)",
    "bn": "Bengali (বাংলা)",
    "ta": "Tamil (தமிழ்)",
    
    # 中东语言
    "ar": "Arabic (العربية)",
    "fa": "Persian (فارسی)",
    "he": "Hebrew (עברית)",
    
    # 其他语言
    "sw": "Swahili (Kiswahili)",
    "el": "Greek (Ελληνικά)",
    "cs": "Czech (Čeština)",
    "hu": "Hungarian (Magyar)",
    "ro": "Romanian (Română)",
    "uk": "Ukrainian (Українська)",
    "fi": "Finnish (Suomi)",
    "sv": "Swedish (Svenska)",
    "da": "Danish (Dansk)",
    "no": "Norwegian (Norsk)",
}
# ========================================================================

SYSTEM_PROMPT = """
You are a professional translator.
Return ONLY valid JSON with the EXACT top-level key "segments".
Example:
{
  "segments": [
    {"id": 0, "text": "Hello"},
    {"id": 1, "text": "World"}
  ]
}
""".strip()

# ----------------------- OpenAI 客户端封装 -------------------------------
if _new_sdk:
    _client = OpenAI(api_key=API_KEY, base_url=BASE_URL)

    def chat_complete(messages, **kw):
        return _client.chat.completions.create(model=MODEL,
                                               messages=messages, **kw)
else:
    openai.api_key  = API_KEY
    openai.api_base = BASE_URL

    def chat_complete(messages, **kw):
        return openai.ChatCompletion.create(model=MODEL,
                                            messages=messages, **kw)
# -------------------------------------------------------------------------

enc = tiktoken.encoding_for_model("gpt-4")  # token 估算器


def _translate_segments(segments: List[Dict], target_lang: str) -> Dict[int, str]:
    """Translate list of {'id','text'} dicts; return {id: translated_text}."""
    # ---- batch split ----------------------------------------------------
    batches, batch, tokens = [], [], 0
    for seg in segments:
        tok = len(enc.encode(seg["text"])) or 1
        if tokens + tok > MAX_BATCH_TOKENS and batch:
            batches.append(batch); batch, tokens = [], 0
        batch.append(seg); tokens += tok
    if batch:
        batches.append(batch)

    out: Dict[int, str] = {}
    # ---- send -----------------------------------------------------------
    for b in tqdm(batches, desc="Translating via DeepSeek", unit="batch"):
        prompt = [
            {"role": "system", "content": SYSTEM_PROMPT},
            {"role": "user",   "content": json.dumps({
                "meta": {"source_lang": "auto", "target_lang": target_lang},
                "segments": b}, ensure_ascii=False)}
        ]
        rsp = chat_complete(prompt,
                            response_format={"type": "json_object"},
                            temperature=TEMPERATURE,
                            timeout=REQUEST_TIMEOUT)

        data = json.loads(rsp.choices[0].message.content)
        key = "segments" if "segments" in data else (
              "translation" if "translation" in data else None)
        if key is None:
            raise ValueError(f"Unexpected JSON keys: {list(data.keys())}")
        for item in data[key]:
            out[item["id"]] = item["text"]
        time.sleep(0.2)
    return out


# -------------------------- Word ----------------------------------------
def translate_docx(path: Path, target_lang: str) -> Path:
    doc = Document(path)
    segments, mapping, idx = [], [], 0
    for p in doc.paragraphs:
        for r in p.runs:
            if r.text.strip():
                segments.append({"id": idx, "text": r.text})
                mapping.append(r); idx += 1
    if not segments:
        print("[WARN] No translatable text found in", path)
        return path
    translated = _translate_segments(segments, target_lang)
    for i, run in enumerate(mapping):
        run.text = translated.get(i, run.text)
    out = path.with_stem(f"{path.stem}.{target_lang}")
    doc.save(out); return out


# -------------------------- Excel ---------------------------------------
def translate_xlsx(path: Path, target_lang: str) -> Path:
    wb = load_workbook(path, keep_vba=True)  # 保留VBA
    segments, mapping, idx = [], [], 0
    for ws in wb.worksheets:
        for row in ws.iter_rows():
            for c in row:
                if isinstance(c.value, str) and c.value.strip():
                    segments.append({"id": idx, "text": c.value})
                    mapping.append(c); idx += 1
    if not segments:
        print("[WARN] No translatable text found in", path)
        return path
    translated = _translate_segments(segments, target_lang)
    
    # 创建新工作簿并复制所有工作表
    out = path.with_stem(f"{path.stem}.{target_lang}")
    if out.exists():
        out.unlink()  # 如果目标文件已存在，先删除
        
    # 保存工作簿前更新翻译内容
    for i, cell in enumerate(mapping):
        cell.value = translated.get(i, cell.value)
    
    # 保存时保留所有格式
    wb.save(out)
    return out


def get_supported_languages():
    """返回支持的语言代码和名称的格式化字符串"""
    lines = []
    for code, name in sorted(LANGUAGE_CODES.items()):
        lines.append(f"  {code:<6} - {name}")
    return "\n".join(lines)

def validate_language_code(code: str) -> str:
    """验证语言代码是否支持，返回标准化的语言代码"""
    code = code.lower()
    if code in LANGUAGE_CODES:
        return code
    # 处理一些常见的别名
    aliases = {
        "cn": "zh",
        "zh-cn": "zh",
        "zht": "zh-tw",
        "jpn": "ja",
        "kor": "ko",
        "eng": "en",
        "fra": "fr",
        "deu": "de",
        "spa": "es",
    }
    if code in aliases:
        return aliases[code]
    
    raise ValueError(
        f"Unsupported language code: {code}\n"
        f"Supported languages:\n{get_supported_languages()}"
    )

# --------------------------- CLI ----------------------------------------
def main():
    pa = argparse.ArgumentParser(
        description="Translate .docx / .xlsx via DeepSeek while keeping formatting",
        formatter_class=argparse.RawDescriptionHelpFormatter)
    pa.add_argument("file", type=Path)
    pa.add_argument("-l", "--target", required=True,
                    help="Target language code (use -h to see supported languages)")
    pa.add_argument("--list-languages", action="store_true",
                    help="List all supported languages and exit")
    args = pa.parse_args()

    if args.list_languages:
        print("\nSupported languages:")
        print(get_supported_languages())
        sys.exit(0)

    if not args.file.exists():
        print("[ERR] file not found:", args.file)
        sys.exit(1)

    try:
        # 验证并标准化语言代码
        target_lang = validate_language_code(args.target)
        
        if args.file.suffix.lower() == ".docx":
            out = translate_docx(args.file, target_lang)
        elif args.file.suffix.lower() == ".xlsx":
            out = translate_xlsx(args.file, target_lang)
        else:
            print("[ERR] Use .docx or .xlsx")
            sys.exit(1)
    except ValueError as e:
        print("\n[ERR]", str(e))
        sys.exit(1)
    except Exception as e:
        print("\n[FAIL]", e)
        sys.exit(1)

    print("\n✅ Done! Output:", out)


if __name__ == "__main__":
    main()
